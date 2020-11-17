def htmlToStr(s):
    if not s:
        return ""

    mapping = {
        "&#039;": "'",
        "&quot;": "\"",
        "&amp;": "&",
        "&nbsp;": " "
    }
    
    for m in mapping:
        if m in s:
            s = s.replace(m, mapping[m])
    return s

def indicateModifications(html_file, xlsx_file):
    import openpyxl
    import xml.etree.ElementTree as ET

    output_xlsx_filename = xlsx_file.split(".")[0] + "_output.xlsx"
    lines = set()
    wb = openpyxl.load_workbook(xlsx_file)
    root = ET.parse(html_file).getroot()
    
    for i in root.iter('linecomp'):
        if i.attrib['status'] == "different" or i.attrib['status'] =="leftorphan":
            original_text = htmlToStr(i[0].text)
            if original_text and not original_text.isspace() and not len(original_text) == 0: 
                lines.add(original_text)
            
    for line in lines:
        print(line)
            
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]

        i = 4
        while True:
            field = sheet[f"C{i}"].value
            if field:
                field = str(field)
                for line in lines:
                    if line in field:
                        sheet[f"C{i}"].fill = openpyxl.styles.PatternFill(fgColor="FFC7CE", fill_type="solid")
                        break
            else:
                break

            i += 1
        
        wb.save(output_xlsx_filename)
